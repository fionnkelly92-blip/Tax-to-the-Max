# excel_generator.py

import os

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

from engine import (
    COMPANIES,
    SCENARIOS,
    TOTAL_REVENUE,
    expected_loss_table,
    project_revenue,
    run_monte_carlo,
)

# ----------------------------------------------------
# STYLE HELPERS
# ----------------------------------------------------

HEADER_FONT = Font(bold=True)
CENTERED = Alignment(horizontal="center", vertical="center")
HEADER_FILL = PatternFill(start_color="DDDDDD", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


def _auto_fit_columns(ws):
    """Simple auto-width for all used columns in a sheet."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        # small padding
        ws.column_dimensions[col_letter].width = max_len + 2


# ----------------------------------------------------
# MAIN EXCEL GENERATION FUNCTION
# ----------------------------------------------------

def generate_excel(
    output_file: str = "Tax_Shock_Model.xlsx",
    scenario: str = "Medium",
    runs: int = 1000,
    growth: float = 0.03,
):
    """
    Build a full Excel workbook for the 'Taxed To The Max' project:
    - Data Model
    - Controls
    - Dashboard (with charts)
    - Projections
    - Monte Carlo simulation
    """

    # ====================================================
    # 1. CREATE WORKBOOK
    # ====================================================
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Model"

    # ====================================================
    # 2. DATA MODEL SHEET
    # ====================================================
    headers = [
        "Company", "Sector", "Tax Contribution",
        "Base Exit Prob", "Scenario",
        "Adjusted Exit Prob", "Loss If Leaves",
        "Expected Loss"
    ]

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTERED
        cell.border = THIN_BORDER

    # Get expected-loss table from engine
    df = expected_loss_table(scenario)

    start_row = 2
    for i, (_, row) in enumerate(df.iterrows(), start=start_row):
        ws.cell(row=i, column=1, value=row["Company"])
        ws.cell(row=i, column=2, value=row["Sector"])
        ws.cell(row=i, column=3, value=row["Tax Contribution"])
        ws.cell(row=i, column=4, value=row["Base Prob"])
        ws.cell(row=i, column=5, value=scenario)
        ws.cell(row=i, column=6, value=row["Adj Prob"])
        ws.cell(row=i, column=7, value=row["Tax Contribution"])  # loss if leaves
        ws.cell(row=i, column=8, value=row["Expected Loss"])

        # borders
        for c in range(1, 9):
            ws.cell(row=i, column=c).border = THIN_BORDER

    # Number formatting
    n_rows = len(df)
    for r in range(2, 2 + n_rows):
        ws.cell(row=r, column=3).number_format = "#,##0"   # Tax Contribution
        ws.cell(row=r, column=4).number_format = "0.00%"   # Base Prob
        ws.cell(row=r, column=6).number_format = "0.00%"   # Adj Prob
        ws.cell(row=r, column=7).number_format = "#,##0"   # Loss If Leaves
        ws.cell(row=r, column=8).number_format = "#,##0"   # Expected Loss

    # ====================================================
    # 3. SUMMARY ROWS
    # ====================================================
    sum_row = n_rows + 4

    total_expected_loss = float(df["Expected Loss"].sum())
    post_shock_revenue = float(TOTAL_REVENUE - total_expected_loss)

    ws.cell(row=sum_row, column=1, value="Total Revenue")
    ws.cell(row=sum_row, column=2, value=TOTAL_REVENUE)

    ws.cell(row=sum_row + 1, column=1, value="Total Expected Loss")
    ws.cell(row=sum_row + 1, column=2, value=total_expected_loss)

    ws.cell(row=sum_row + 2, column=1, value="Post-Shock Revenue")
    ws.cell(row=sum_row + 2, column=2, value=post_shock_revenue)

    for r in range(sum_row, sum_row + 3):
        ws.cell(row=r, column=1).font = HEADER_FONT
        ws.cell(row=r, column=2).number_format = "#,##0"

    ws.freeze_panes = "A2"
    _auto_fit_columns(ws)

    # ====================================================
    # 4. CONTROLS SHEET
    # ====================================================
    ws2 = wb.create_sheet("Controls")

    ws2["A1"] = "Selected Scenario"
    ws2["B1"] = scenario

    dv = DataValidation(type="list", formula1='"Mild,Medium,Severe"')
    ws2.add_data_validation(dv)
    dv.add(ws2["B1"])

    ws2["A3"] = "Monte Carlo Runs"
    ws2["B3"] = runs

    ws2["A5"] = "Growth Rate"
    ws2["B5"] = growth

    ws2["A7"] = "Total Revenue (Master)"
    ws2["B7"] = TOTAL_REVENUE

    for cell in ["A1", "A3", "A5", "A7"]:
        ws2[cell].font = HEADER_FONT

    ws2["B7"].number_format = "#,##0"
    _auto_fit_columns(ws2)

    # NOTE: Changing B1 in Excel will NOT auto-recalculate the model.
    # To regenerate for another scenario, re-run this Python script
    # with scenario="Mild" or "Severe".

    # ====================================================
    # 5. DASHBOARD SHEET
    # ====================================================
    ws3 = wb.create_sheet("Dashboard")
    ws3["A1"] = "SUMMARY KPI"
    ws3["A1"].font = Font(bold=True, size=14)

    ws3["A3"] = "Scenario"
    ws3["B3"] = scenario

    ws3["A4"] = "Total Expected Loss"
    ws3["B4"] = total_expected_loss

    ws3["A5"] = "Post-Shock Revenue"
    ws3["B5"] = post_shock_revenue

    ws3["B4"].number_format = "#,##0"
    ws3["B5"].number_format = "#,##0"

    for cell in ["A3", "A4", "A5"]:
        ws3[cell].font = HEADER_FONT

    # ====================================================
    # 6. DASHBOARD CHARTS (matplotlib → images → Excel)
    # ====================================================
    # Ensure we save images beside the Excel file
    img_dir = os.path.dirname(os.path.abspath(output_file))
    if not os.path.exists(img_dir):
        os.makedirs(img_dir, exist_ok=True)

    contrib_plot = os.path.join(img_dir, "contrib.png")
    plt.figure(figsize=(6, 4))
    plt.bar(df["Company"], df["Tax Contribution"])
    plt.xticks(rotation=45, ha="right")
    plt.title("Tax Contribution by Company")
    plt.tight_layout()
    plt.savefig(contrib_plot)
    plt.close()

    ws3.add_image(XLImage(contrib_plot), "A7")

    loss_plot = os.path.join(img_dir, "loss.png")
    plt.figure(figsize=(6, 4))
    plt.bar(df["Company"], df["Expected Loss"])
    plt.xticks(rotation=45, ha="right")
    plt.title("Expected Loss by Company")
    plt.tight_layout()
    plt.savefig(loss_plot)
    plt.close()

    ws3.add_image(XLImage(loss_plot), "A25")
    _auto_fit_columns(ws3)

    # ====================================================
    # 7. PROJECTIONS SHEET
    # ====================================================
    ws4 = wb.create_sheet("Projections")
    proj = project_revenue(10, growth)  # 10-year window

    # Header row
    ws4.append(list(proj.columns))
    for c in range(1, len(proj.columns) + 1):
        cell = ws4.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTERED
        cell.border = THIN_BORDER

    # Data rows
    for _, row in proj.iterrows():
        ws4.append(row.tolist())

    # Format revenue column (assumed col 2 is ProjectedRevenue)
    for row_idx in range(2, 2 + len(proj)):
        cell = ws4.cell(row=row_idx, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "#,##0"

    for row in ws4.iter_rows(min_row=2, max_row=1 + len(proj), min_col=1, max_col=len(proj.columns)):
        for cell in row:
            cell.border = THIN_BORDER

    proj_plot = os.path.join(img_dir, "projection.png")
    plt.figure(figsize=(6, 4))
    plt.plot(proj["Year"], proj["ProjectedRevenue"], marker="o")
    plt.title("Revenue Projection")
    plt.xlabel("Year")
    plt.ylabel("Projected Revenue")
    plt.tight_layout()
    plt.savefig(proj_plot)
    plt.close()

    ws4.add_image(XLImage(proj_plot), "D2")
    _auto_fit_columns(ws4)

    # ====================================================
    # 8. MONTE CARLO SIMULATION SHEET
    # ====================================================
    ws5 = wb.create_sheet("Monte Carlo")

    mc = run_monte_carlo(runs=runs, scenario=scenario)
    mc_series = mc["PostShockRevenue"]  # series-like

    ws5.append(["Run", "PostShockRevenue"])
    ws5["A1"].font = HEADER_FONT
    ws5["B1"].font = HEADER_FONT
    ws5["A1"].fill = HEADER_FILL
    ws5["B1"].fill = HEADER_FILL

    for i, v in enumerate(mc_series, start=2):
        ws5.cell(row=i, column=1, value=i - 1)
        ws5.cell(row=i, column=2, value=float(v))
        ws5.cell(row=i, column=2).number_format = "#,##0"
        ws5.cell(row=i, column=1).border = THIN_BORDER
        ws5.cell(row=i, column=2).border = THIN_BORDER

    mc_plot = os.path.join(img_dir, "mc_hist.png")
    plt.figure(figsize=(6, 4))
    plt.hist(mc_series, bins=30)
    plt.title("Monte Carlo: Post-Shock Revenue")
    plt.xlabel("Post-Shock Revenue")
    plt.ylabel("Frequency")
    plt.tight_layout()
    plt.savefig(mc_plot)
    plt.close()

    ws5.add_image(XLImage(mc_plot), "D2")
    _auto_fit_columns(ws5)

    # ====================================================
    # 9. SAVE EXCEL FILE
    # ====================================================
    wb.save(output_file)
    print(f"Excel model generated: {output_file}")


if __name__ == "__main__":
    # You can change these to regenerate different versions
    generate_excel(
        output_file="Tax_Shock_Model.xlsx",
        scenario="Medium",   # "Mild", "Medium", "Severe"
        runs=1000,
        growth=0.03,
    )
